#!/usr/bin/env python
# coding: utf-8

# Execute the script to save the DataFrame to an Excel file
import anthropic
import openpyxl

# Set the Claude 3.5 Sonnet API key
client = anthropic.Anthropic(
    api_key="Your api_key"
)

# Read an Excel file
workbook = openpyxl.load_workbook("Path to your file")
sheet = workbook.active

# Apply the range of the row and column
start_row = "number"
end_row = "number"
question_column = "number"
answer_column = "number"

# Repeat the execution of the command 
for row in range(start_row, end_row + 1):
    question = sheet.cell(row=row, column=question_column).value
    if question:
        # Answer the question using the Claude 3.5 Sonnet API
        response = client.messages.create(
            model="claude-3-5-sonnet-20240620",
            max_tokens=4096,
            messages=[
                {"role": "user", "content": question}
            ]
        )
        answer = response.content[0].text
        # Write an answer next to the cell
        sheet.cell(row=row, column=answer_column).value = answer

# Save the edited Excel file
workbook.save("Path to your file")