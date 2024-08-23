#!/usr/bin/env python
# coding: utf-8

# Execute the script to save the DataFrame to an Excel file
import google.generativeai as genai
import openpyxl

# Set the Gemini 1.5 Pro API key
genai.configure(api_key="Your api_key")

# Set the safety settings
safety_settings = [
    {
        "category": "HARM_CATEGORY_HARASSMENT",
        "threshold": "BLOCK_NONE"
    },
    {
        "category": "HARM_CATEGORY_HATE_SPEECH",
        "threshold": "BLOCK_NONE"
    },
    {
        "category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
        "threshold": "BLOCK_NONE"
    },
    {
        "category": "HARM_CATEGORY_DANGEROUS_CONTENT",
        "threshold": "BLOCK_NONE"
    }
]

# Set the model
model = genai.GenerativeModel(model_name="gemini-1.5-pro-latest", safety_settings=safety_settings)

# Read an Excel file
workbook = openpyxl.load_workbook("Path to your file")
sheet = workbook.active

# Apply the range of the row and column
start_row = "number"
end_row = "number"
question_column = "number"
answer_column = "number"

# Repeat the execution of the command 
for row in range(start_row, end_row + 1):
    question = sheet.cell(row=row, column=question_column).value
    if question:
        # Answer the question using the Gemini 1.5 Pro API
        response = model.generate_content(question)
        answer = response.text
        # Write an answer next to the cell
        sheet.cell(row=row, column=answer_column).value = answer

# Save the edited Excel file
workbook.save("Path to your file")